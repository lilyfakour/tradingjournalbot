"""Offline smoke test for the reply-keyboard /trade flow (no Telegram needed).

Run:  .\\.venv\\Scripts\\python.exe smoke_test.py
Uses a throwaway SQLite database and screenshot folder in a temp directory;
safe to run anytime.
"""

import asyncio
import os
import sqlite3
import sys
import tempfile
from datetime import date, datetime as _dt
from pathlib import Path
from types import SimpleNamespace

# Emoji check labels must survive Windows pipes/consoles (cp1252).
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

_TMP = Path(tempfile.mkdtemp(prefix="journal_smoke_"))
os.environ["JOURNAL_DB"] = str(_TMP / "smoke.db")

import db  # noqa: E402
import export  # noqa: E402
import journal  # noqa: E402

journal.SCREENSHOT_DIR = _TMP / "shots"
export.EXPORT_DIR = _TMP / "exports"

from telegram import (  # noqa: E402
    CallbackQuery,
    Chat,
    InlineKeyboardMarkup,
    Message,
    MessageEntity,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
    User,
)
from telegram.ext import CallbackQueryHandler  # noqa: E402


class FakeChat:
    id = 7

    def __init__(self, log):
        self._log = log

    async def send_message(self, text, reply_markup=None, **kwargs):
        self._log.append(("send", text, reply_markup))
        # A real send returns the sent message (stats() stores its id).
        return SimpleNamespace(message_id=100 + len(self._log))

    async def send_action(self, action):
        self._log.append(("action", action, None))

    async def send_photo(self, photo, caption=None, **kwargs):
        self._log.append(("photo", caption, None))


class FakeTgFile:
    def __init__(self, saved):
        self._saved = saved

    async def download_to_drive(self, path):
        self._saved.append(Path(path))
        Path(path).write_bytes(b"fake image bytes")
        return str(path)


class FakePhotoSize:
    def __init__(self, saved):
        self._saved = saved

    async def get_file(self):
        return FakeTgFile(self._saved)


class FakeMessage:
    def __init__(self, text=None, photo=None, log=None):
        self.text = text
        self.photo = photo or []
        self.document = None
        self._log = log

    async def reply_text(self, text, reply_markup=None, **kwargs):
        self._log.append(("reply", text, reply_markup))


class FakeUpdate:
    def __init__(self):
        self.sent = []
        self.effective_chat = FakeChat(self.sent)
        self.message = None

    @property
    def effective_message(self):
        return self.message

    def text(self, value):
        self.message = FakeMessage(text=value, log=self.sent)
        return self

    def photo(self):
        self.message = FakeMessage(photo=[FakePhotoSize(self.sent)], log=self.sent)
        return self


class FakeContext:
    def __init__(self, args=None):
        self.user_data = {}
        self.args = args or []
        self.bot = None  # stats callbacks send/refresh via context.bot


class FakeBot:
    """Just enough bot for the stats inline-button callbacks."""

    def __init__(self, log):
        self._log = log
        self.sent_docs = []
        self._next_id = 500

    async def send_message(self, chat_id, text, reply_markup=None, **kwargs):
        self._log.append(("bot-send", text, reply_markup))
        message = SimpleNamespace(message_id=self._next_id)
        self._next_id += 1
        return message

    async def edit_message_text(
        self, text, chat_id=None, message_id=None, reply_markup=None, **kwargs
    ):
        self._log.append(("bot-edit", text, reply_markup))
        return SimpleNamespace(message_id=message_id)

    async def send_document(
        self, chat_id, document, filename=None, caption=None, **kwargs
    ):
        self.sent_docs.append((filename, caption, document.read()))

    async def send_chat_action(self, chat_id, action, **kwargs):
        self._log.append(("action", action, None))


def _labels(markup):
    return [btn.text for row in markup.keyboard for btn in row]


def _inline_flat(markup):
    """(text, callback_data) pairs of an InlineKeyboardMarkup."""
    return [
        (b.text, b.callback_data) for row in markup.inline_keyboard for b in row
    ]


def _home_labels():
    return [
        "📈 بستن معامله",
        "🟢 ثبت معامله باز",
        "🟢 معاملات باز",
        "📊 آمار",
        "🕘 معاملات اخیر",
        "📥 اکسل",
        "🏠 منو",
    ]


def _last_markup(upd):
    """Last non-None reply markup (menu re-sends carry none)."""
    for entry in reversed(upd.sent):
        if entry[2] is not None:
            return entry[2]
    return None


def _markup_with(upd, marker):
    """Last reply markup containing a button with *marker* in its label."""
    for entry in reversed(upd.sent):
        markup = entry[2]
        if markup is not None and any(
            marker in btn.text for row in markup.keyboard for btn in row
        ):
            return markup
    return None


def _reply_text(upd):
    """Last reply text, skipping trailing main-menu re-sends."""
    for entry in reversed(upd.sent):
        if entry[1] != journal.MENU_TEXT:
            return entry[1]
    return ""


async def main() -> int:
    failures = []

    def check(cond, label):
        print(("PASS  " if cond else "FAIL  ") + label)
        if not cond:
            failures.append(label)

    db.init_db()
    cols = {
        r[1]
        for r in sqlite3.connect(os.environ["JOURNAL_DB"]).execute(
            "PRAGMA table_info(trades)"
        )
    }
    check(
        "timeframe" in cols
        and "mood" in cols
        and "screenshot" in cols
        and "market" in cols
        and "leverage" in cols
        and "risk_percent" in cols
        and "take_profit" in cols
        and "stop_loss" in cols
        and "hit" in cols
        and "screenshot_after" in cols
        and "entry_time" in cols
        and "exit_time" in cols
        and "entry_reason" in cols
        and "exit_reason" in cols
        and "exit_photos" in cols
        and "source" in cols,
        "new database has all columns (open-flow + legacy)",
    )
    open_cols = {
        r[1]
        for r in sqlite3.connect(os.environ["JOURNAL_DB"]).execute(
            "PRAGMA table_info(open_trades)"
        )
    }
    check(
        {
            "symbol", "direction", "market", "timeframe", "reason",
            "screenshot", "trade_date", "entry_time", "risk_percent",
            "entry_price", "take_profit", "stop_loss",
        }
        <= open_cols,
        "open_trades table has every open-questionnaire column",
    )

    # --- migration: an old-schema database gains the new columns -------------
    old_db = _TMP / "old.db"
    conn = sqlite3.connect(old_db)
    conn.execute(
        "CREATE TABLE trades ("
        " id INTEGER PRIMARY KEY AUTOINCREMENT,"
        " symbol TEXT NOT NULL,"
        " direction TEXT NOT NULL,"
        " entry_price REAL NOT NULL,"
        " exit_price REAL NOT NULL,"
        " size REAL NOT NULL,"
        " pnl REAL NOT NULL,"
        " trade_date TEXT NOT NULL,"
        " notes TEXT NOT NULL DEFAULT '',"
        " created_at TEXT NOT NULL DEFAULT (datetime('now')))"
    )
    conn.commit()
    conn.close()
    real_db_path = db.DB_PATH
    db.DB_PATH = old_db
    db.init_db()
    cols = {r[1] for r in sqlite3.connect(old_db).execute("PRAGMA table_info(trades)")}
    check(
        "timeframe" in cols and "mood" in cols and "screenshot" in cols
        and "market" in cols and "hit" in cols,
        "init_db migrates old databases",
    )
    notnull = {
        r[1]: r[3]
        for r in sqlite3.connect(old_db).execute("PRAGMA table_info(trades)")
    }
    check(
        notnull["pnl"] == 0 and notnull["size"] == 0,
        "init_db relaxes pnl/size to nullable (open-flow closes have no margin)",
    )
    check(
        "entry_time" in cols and "exit_photos" in cols and "source" in cols,
        "init_db adds the open-flow columns to old databases",
    )
    db.DB_PATH = real_db_path

    # --- timeframe parsing ----------------------------------------------------
    check(
        journal._parse_timeframe("5min") == "5m"
        and journal._parse_timeframe("1day") == "1D"
        and journal._parse_timeframe("daily") == "1D"
        and journal._parse_timeframe("weekly") == "1W"
        and journal._parse_timeframe("M") == "1M"
        and journal._parse_timeframe("45m") == "45m"
        and journal._parse_timeframe("banana") is None,
        "timeframe parsing / normalization",
    )

    # --- skip-button tokens & P&L/ROI formatters --------------------------------
    check(
        "⏭ رد کردن" in journal._SKIP_TOKENS
        and "⏭ skip" in journal._SKIP_TOKENS
        and "⏭ بدون اسکرین‌شات" in journal._SKIP_SHOT_TOKENS
        and "⏭ بدون دلیل" in journal._SKIP_NOTES_TOKENS,
        "⏭ button labels match the skip tokens (mood/hour bugfix)",
    )
    check(
        journal._fmt_pnl(2.0) == "+$2.00"
        and journal._fmt_pnl(-0.02) == "-$0.02"
        and journal._fmt_pnl(0) == "+$0.00"
        and journal._fmt_roi(12.5) == "+12.50%"
        and journal._fmt_roi(None) == "-",
        "P&L / ROI formatting helpers",
    )

    ctx = FakeContext()
    upd = FakeUpdate()

    # --- happy path: reply-keyboard buttons -----------------------------------
    state = await journal.trade_start(upd.text("/trade"), ctx)
    check(state == journal.MARKET, "/trade starts at MARKET")
    market_labels = _labels(_last_markup(upd))
    check(
        "🪙 کریپتو" in market_labels and "💵 فارکس" in market_labels,
        "market buttons present",
    )

    state = await journal.ask_market(upd.text("🪙 کریپتو"), ctx)
    check(
        state == journal.SYMBOL and ctx.user_data["market"] == "crypto",
        "market choice -> SYMBOL",
    )
    check(
        isinstance(_last_markup(upd), ReplyKeyboardRemove),
        "typing prompt hides the keyboard (menu only at the end)",
    )

    state = await journal.ask_symbol(upd.text("eurusd"), ctx)
    check(
        state == journal.DIRECTION and ctx.user_data["symbol"] == "EURUSD",
        "symbol via text -> DIRECTION",
    )
    markup = _last_markup(upd)
    check(
        isinstance(markup, ReplyKeyboardMarkup),
        "direction prompt uses a reply keyboard (under the screen)",
    )
    labels = _labels(markup)
    check(
        "📈 Long" in labels and "📉 Short" in labels and "✖️ لغو" in labels,
        "direction buttons present (English Long/Short)",
    )

    state = await journal.ask_direction(upd.text("📈 خرید"), ctx)
    check(
        state == journal.LEVERAGE and ctx.user_data["direction"] == "long",
        "direction -> LEVERAGE",
    )
    lev_labels = _labels(_last_markup(upd))
    check(
        all(
            lbl in lev_labels
            for lbl in ("×2", "×10", "×100", "⏭ بدون اهرم", "✖️ لغو")
        ),
        "leverage buttons present",
    )

    state = await journal.ask_leverage(upd.text("10x"), ctx)
    check(
        state == journal.TIMEFRAME and ctx.user_data["leverage"] == 10,
        "typed '10x' -> TIMEFRAME with leverage 10",
    )
    tf_labels = _labels(upd.sent[-1][2])
    check(
        all(
            t in tf_labels
            for t in ("1m", "5m", "15m", "30m", "1h", "4h", "1D", "1W", "1M")
        ),
        "timeframe buttons present",
    )

    state = await journal.ask_timeframe(upd.text("1h"), ctx)
    check(
        state == journal.ENTRY and ctx.user_data["timeframe"] == "1h",
        "typed '1h' -> ENTRY",
    )
    check(
        isinstance(_last_markup(upd), ReplyKeyboardRemove),
        "typing prompts hide the keyboard (menu only at the end)",
    )

    await journal.ask_entry(upd.text("100"), ctx)
    state = await journal.ask_take_profit(upd.text("110"), ctx)
    check(state == journal.STOP_LOSS, "entry + TP -> STOP_LOSS")
    state = await journal.ask_stop_loss(upd.text("95"), ctx)
    check(
        state == journal.RESULT and ctx.user_data["stop_loss"] == 95,
        "SL -> RESULT",
    )
    state = await journal.ask_result(upd.text("✅ Win"), ctx)
    check(
        state == journal.MARGIN
        and ctx.user_data["hit"] == "win"
        and ctx.user_data["exit_price"] == 110,
        "Win -> MARGIN with exit = TP",
    )
    state = await journal.ask_margin(upd.text("2"), ctx)
    check(
        state == journal.RISK and ctx.user_data["size"] == 2,
        "margin stored -> RISK",
    )
    state = await journal.ask_risk(upd.text("1%"), ctx)
    check(
        state == journal.TRADE_DATE and ctx.user_data["risk_percent"] == 1,
        "risk stored -> TRADE_DATE (P&L auto-calculated, no question)",
    )

    state = await journal.ask_trade_date(upd.text("📅 امروز"), ctx)
    check(
        state == journal.TRADE_HOUR
        and ctx.user_data["trade_date"] == date.today().isoformat(),
        "Today button -> TRADE_HOUR",
    )
    hour_labels = _labels(upd.sent[-1][2])
    check(
        all(
            h in hour_labels
            for h in ("00", "03", "06", "09", "12", "15", "18", "21")
            + ("⏭ رد کردن", "✖️ لغو")
        ),
        "hour buttons present",
    )
    state = await journal.ask_trade_hour(upd.text("14:30"), ctx)
    check(
        state == journal.NOTES
        and ctx.user_data["trade_date"] == f"{date.today().isoformat()} 14:30",
        "typed HH:MM -> NOTES with time appended",
    )

    state = await journal.ask_notes(upd.text("⏭ بدون دلیل"), ctx)
    check(
        state == journal.MOOD and ctx.user_data["notes"] == "",
        "Skip notes button -> MOOD",
    )
    mood_labels = _labels(upd.sent[-1][2])
    check(
        all(
            m in mood_labels
            for m in (
                "آرام",
                "مطمئن",
                "مضطرب",
                "طمع",
                "FOMO",
                "انتقامی",
                "⏭ رد کردن",
                "✖️ لغو",
            )
        ),
        "mood buttons present (plain professional labels, no faces)",
    )
    state = await journal.ask_mood(upd.text("مضطرب"), ctx)
    check(
        state == journal.SCREENSHOT and ctx.user_data["mood"] == "anxious",
        "Anxious mood button -> SCREENSHOT",
    )

    state = await journal.ask_screenshot(upd.photo(), ctx)
    shot_name = ctx.user_data.get("screenshot")
    check(
        state == journal.SCREENSHOT_AFTER and shot_name,
        "photo (before) stored -> SCREENSHOT_AFTER",
    )
    state = await journal.ask_screenshot_after_text(upd.text("-"), ctx)
    check(state == journal.CONFIRM, "after-shot skipped -> CONFIRM")
    summary = upd.sent[-1][1]
    check(
        "TF·Lev" in summary and "1h" in summary and "10x" in summary,
        "summary shows timeframe and leverage",
    )
    check("Market" in summary and "🪙 کریپتو" in summary, "summary shows market")
    check(
        "Result" in summary and "🟢" in summary and "Win" in summary,
        "summary shows the result with emoji",
    )
    check("Margin" in summary, "summary shows margin")
    check("Risk" in summary and "1%" in summary, "summary shows risk")
    check(
        "Mood" in summary and "مضطرب" in summary,
        "summary shows mood",
    )
    check("Shots" in summary and "قبل" in summary, "summary shows before screenshot")
    check(
        "P&L" in summary and "+$2.00" in summary,
        "summary shows signed $ P&L",
    )
    check(
        "ROI" in summary and "+100.00%" in summary,
        "summary shows per-trade ROI",
    )
    check(
        "__" not in summary and "•" in summary and "\n\n" in summary,
        "summary is airy: bullets, blank lines, no literal underscores",
    )

    state = await journal.save_trade(upd.text("✅ ذخیره"), ctx)
    check(
        state == journal.ConversationHandler.END,
        "Save button -> END",
    )
    check(not ctx.user_data, "draft cleared after save")
    check(
        _labels(_last_markup(upd)) == _home_labels(),
        "after save the persistent main-menu bar is shown again",
    )
    conf = upd.sent[-1][1]
    check(
        "ذخیره شد" in conf and "__" not in conf and "\n\n" in conf,
        "save confirmation is airy and bold, no literal underscores",
    )

    # --- second trade: forex flow, typed fallbacks, manual P&L ----------------
    ctx2 = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx2)
    state = await journal.ask_market(upd.text("💵 فارکس"), ctx2)
    check(
        state == journal.SYMBOL and ctx2.user_data["market"] == "forex",
        "forex market -> SYMBOL",
    )
    await journal.ask_symbol(upd.text("BTCUSD"), ctx2)
    await journal.ask_direction(upd.text("S"), ctx2)
    state = await journal.ask_leverage(upd.text("×10"), ctx2)
    check(state == journal.TIMEFRAME and ctx2.user_data["leverage"] == 10, "×10 button parsed")
    state = await journal.ask_timeframe(upd.text("banana"), ctx2)
    check(state == journal.TIMEFRAME, "invalid timeframe reprompts")
    state = await journal.ask_timeframe(upd.text("daily"), ctx2)
    check(
        state == journal.ENTRY and ctx2.user_data["timeframe"] == "1D",
        "typed 'daily' normalized to 1D",
    )
    await journal.ask_entry(upd.text("50000"), ctx2)
    await journal.ask_take_profit(upd.text("49000"), ctx2)
    await journal.ask_stop_loss(upd.text("51000"), ctx2)
    state = await journal.ask_result(upd.text("❌ Loss"), ctx2)
    check(
        state == journal.MARGIN
        and ctx2.user_data["hit"] == "lose"
        and ctx2.user_data["exit_price"] == 51000,
        "Loss stores exit = stop loss",
    )
    check("USD" in upd.sent[-1][1], "forex margin prompt mentions dollars")
    state = await journal.ask_margin(upd.text("0.1"), ctx2)
    check(state == journal.RISK, "margin stored -> RISK")
    state = await journal.ask_risk(upd.text("⏭ بدون درصد"), ctx2)
    check(
        state == journal.TRADE_DATE and "risk_percent" not in ctx2.user_data,
        "risk skipped -> TRADE_DATE",
    )
    await journal.ask_trade_date(upd.text("-"), ctx2)
    await journal.ask_trade_hour(upd.text("⏭ رد کردن"), ctx2)
    await journal.ask_notes(upd.text("-"), ctx2)
    await journal.ask_mood(upd.text("⏭ رد کردن"), ctx2)
    state = await journal.ask_screenshot_text(upd.text("⏭ بدون اسکرین‌شات"), ctx2)
    check(
        state == journal.SCREENSHOT_AFTER,
        "before-shot skipped -> SCREENSHOT_AFTER",
    )
    state = await journal.ask_screenshot_after_text(upd.text("-"), ctx2)
    check(state == journal.CONFIRM, "after-shot skipped -> CONFIRM")
    state = await journal.save_trade(upd.text("y"), ctx2)
    check(state == journal.ConversationHandler.END, "typed 'y' saves -> END")

    rows = db.get_recent(2)
    # Auto P&L for this short: 0.1 margin * 10x * ((50000 - 51000) / 50000)
    check(
        rows[0]["symbol"] == "BTCUSD"
        and rows[0]["direction"] == "short"
        and rows[0]["timeframe"] == "1D"
        and abs(rows[0]["pnl"] - (-0.02)) < 1e-9
        and rows[0]["screenshot"] is None
        and rows[0]["mood"] == ""
        and rows[0]["trade_date"] == date.today().isoformat()
        and rows[0]["market"] == "forex"
        and rows[0]["hit"] == "lose"
        and rows[0]["leverage"] == 10
        and rows[0]["risk_percent"] is None
        and rows[0]["take_profit"] == 49000
        and rows[0]["stop_loss"] == 51000
        and rows[0]["exit_price"] == 51000
        and rows[0]["size"] == 0.1
        and rows[0]["roi"] is not None
        and abs(rows[0]["roi"] - (-20.0)) < 1e-9,
        "second trade saved: forex, Loss, auto P&L + ROI, risk skipped",
    )

    # --- 🕘 Recent: one button per trade, tapping SENDS the detail --------------
    recent_upd = FakeUpdate()
    await journal.recent(recent_upd.text("/recent"), FakeContext())
    panel_text = _reply_text(recent_upd)
    check(
        panel_text.startswith("🕘 ") and "صفحه" in panel_text,
        "/recent sends the paginated panel",
    )
    markup = _last_markup(recent_upd)
    check(
        isinstance(markup, InlineKeyboardMarkup),
        "/recent panel uses inline buttons",
    )
    flat_panel = _inline_flat(markup)
    cbs = [cb for _, cb in flat_panel]
    trade_btns = [(t, cb) for t, cb in flat_panel if cb.startswith("rec:v:")]
    check(
        "rec:home" in cbs and "rec:noop" in cbs
        and any(c.startswith("rec:r:") for c in cbs),
        "panel has range, pager and home callbacks",
    )
    check(
        len(trade_btns) == 2,
        "one inline button per trade in the list",
    )
    check(
        all(" EURUSD" in t or " BTCUSD" in t for t, _ in trade_btns)
        and any("+$2.00" in t for t, _ in trade_btns)
        and any("-$0.02" in t for t, _ in trade_btns),
        "trade buttons show the pair and the signed P&L",
    )

    class FakeQuery:
        def __init__(self, data, upd):
            self.data = data
            self._upd = upd
            self.answers = []

        async def answer(self, text=None):
            self.answers.append(text)

        async def edit_message_text(
            self, text, reply_markup=None, **kwargs
        ):
            self._upd.sent.append(("edit", text, reply_markup))

        async def edit_message_reply_markup(
            self, reply_markup=None, **kwargs
        ):
            self._upd.sent.append(("edit-markup", "", reply_markup))

    class FakeCallbackUpdate:
        def __init__(self, data):
            self.sent = []
            self.effective_chat = FakeChat(self.sent)
            self.message = None
            self.callback_query = FakeQuery(data, self)

    async def _tap_recent(data):
        cb_upd = FakeCallbackUpdate(data)
        await journal.on_recent_callback(cb_upd, FakeContext())
        return cb_upd

    # tapping trade #2 (the BTCUSD short: -$0.02 on 0.1 => -20% ROI) SENDS it
    cb_upd = await _tap_recent(f"rec:v:{rows[0]['id']}")
    detail_kind, detail_text, detail_markup = cb_upd.sent[-1]
    check(
        detail_kind == "send",
        "tapping a trade SENDS the detail as its own message",
    )
    check(
        f"معامله #{rows[0]['id']} — <b>BTCUSD</b>" in detail_text
        and "Short" in detail_text and "📉" in detail_text,
        "sent detail shows id, symbol and the side",
    )
    check(
        "-$0.02" in detail_text
        and "-20.00%" in detail_text
        and "• ورود:" in detail_text
        and "• 🎯 هدف:" in detail_text,
        "sent detail shows signed P&L, ROI and airy bullet fields",
    )
    check(
        detail_text.count("\n\n") >= 3 and "——" not in detail_text,
        "detail is airy (blank lines) with no dash rules",
    )
    flat = _inline_flat(detail_markup)
    check(
        ("🗑 حذف", f"rec:d:{rows[0]['id']}") in flat
        and ("❌ بستن", "rec:close") in flat,
        "sent detail has Delete and Close buttons",
    )
    check(
        all("معامله #" not in text for kind, text, _ in cb_upd.sent
            if kind == "edit"),
        "the detail never edits the panel message in place",
    )
    cb_upd = await _tap_recent("rec:p:1")
    check(
        _reply_text(cb_upd).startswith("🕘 "),
        "◀️/▶️ paging still edits the panel in place",
    )
    check(
        rows[1]["symbol"] == "EURUSD"
        and rows[1]["timeframe"] == "1h"
        and rows[1]["pnl"] == 2.0
        and rows[1]["mood"] == "anxious"
        and rows[1]["screenshot"] == shot_name
        and rows[1]["market"] == "crypto"
        and rows[1]["hit"] == "win"
        and rows[1]["leverage"] == 10
        and rows[1]["risk_percent"] == 1
        and rows[1]["take_profit"] == 110
        and rows[1]["stop_loss"] == 95
        and rows[1]["exit_price"] == 110
        and rows[1]["size"] == 2
        and rows[1]["roi"] is not None
        and abs(rows[1]["roi"] - 100.0) < 1e-9,
        "first trade stored with market/hit/leverage/risk + auto P&L/ROI",
    )

    # --- 🟢 open trades: add questionnaire, panel, callbacks --------------------
    octx = FakeContext()
    state = await journal.open_trade_start(upd.text("/open"), octx)
    check(state == journal.OPEN_MARKET, "/open starts at OPEN_MARKET")
    state = await journal.ask_open_market(upd.text("🪙 کریپتو"), octx)
    check(
        state == journal.OPEN_SYMBOL and octx.user_data["market"] == "crypto",
        "open: market -> OPEN_SYMBOL",
    )
    state = await journal.ask_open_symbol(upd.text("xauusd"), octx)
    check(
        state == journal.OPEN_DIRECTION and octx.user_data["symbol"] == "XAUUSD",
        "open: symbol uppercased -> OPEN_DIRECTION",
    )
    state = await journal.ask_open_direction(upd.text("📈 Long"), octx)
    check(
        state == journal.OPEN_TIMEFRAME and octx.user_data["direction"] == "long",
        "open: direction -> OPEN_TIMEFRAME",
    )
    state = await journal.ask_open_timeframe(upd.text("1h"), octx)
    check(state == journal.OPEN_REASON, "open: timeframe -> OPEN_REASON (order)")
    state = await journal.ask_open_reason(upd.text("broke range high"), octx)
    check(
        state == journal.OPEN_SCREENSHOT
        and octx.user_data["reason"] == "broke range high",
        "open: entry reason -> OPEN_SCREENSHOT",
    )
    state = await journal.ask_open_screenshot(upd.photo(), octx)
    open_shot = octx.user_data.get("screenshot")
    check(
        state == journal.OPEN_TRADE_DATE and open_shot,
        "open: entry screenshot -> OPEN_TRADE_DATE",
    )
    state = await journal.ask_open_trade_date(upd.text("2026-03-01"), octx)
    check(
        state == journal.OPEN_TRADE_HOUR
        and octx.user_data["trade_date"] == "2026-03-01",
        "open: entry date -> OPEN_TRADE_HOUR (separate questions)",
    )
    state = await journal.ask_open_trade_hour(upd.text("10:30"), octx)
    check(
        state == journal.OPEN_RISK and octx.user_data["entry_time"] == "10:30",
        "open: entry time -> OPEN_RISK",
    )
    state = await journal.ask_open_risk(upd.text("1%"), octx)
    check(
        state == journal.OPEN_ENTRY and octx.user_data["risk_percent"] == 1,
        "open: risk -> OPEN_ENTRY",
    )
    state = await journal.ask_open_entry(upd.text("2000"), octx)
    check(state == journal.OPEN_TAKE_PROFIT, "open: entry -> OPEN_TAKE_PROFIT")
    state = await journal.ask_open_take_profit(upd.text("2100"), octx)
    check(state == journal.OPEN_STOP_LOSS, "open: TP -> OPEN_STOP_LOSS")
    state = await journal.ask_open_stop_loss(upd.text("1950"), octx)
    check(state == journal.OPEN_CONFIRM, "open: SL -> OPEN_CONFIRM")
    summary = upd.sent[-1][1]
    check(
        "Market" in summary and "🪙 کریپتو" in summary
        and "Symbol" in summary and "XAUUSD" in summary
        and "TF" in summary and "1h" in summary
        and "Reason" in summary and "broke range high" in summary
        and "Shot" in summary and "📷" in summary
        and "Date" in summary and "2026-03-01 10:30" in summary
        and "Risk" in summary and "1%" in summary
        and "Entry" in summary and "2000" in summary
        and "TP / SL" in summary and "2100" in summary and "1950" in summary
        and "معاملات باز" in summary,
        "open: confirmation summary shows every answer (airy HTML)",
    )
    closed_before = db.count_trades()
    state = await journal.save_open_trade(upd.text("✅ ثبت"), octx)
    check(
        state == journal.ConversationHandler.END and not octx.user_data,
        "open: save -> END, draft cleared",
    )
    open_rows = db.get_open_trades(10)
    check(
        db.count_open_trades() == 1 and open_rows[0]["symbol"] == "XAUUSD",
        "open: trade stored in open_trades",
    )
    xau = open_rows[0]
    check(
        xau["direction"] == "long" and xau["market"] == "crypto"
        and xau["timeframe"] == "1h" and xau["reason"] == "broke range high"
        and xau["screenshot"] == open_shot
        and xau["trade_date"] == "2026-03-01" and xau["entry_time"] == "10:30"
        and xau["risk_percent"] == 1 and xau["entry_price"] == 2000
        and xau["take_profit"] == 2100 and xau["stop_loss"] == 1950,
        "open: every questionnaire answer stored",
    )
    check(
        db.count_trades() == closed_before,
        "adding an open trade does not touch the closed list",
    )

    # --- symbol suggestions: most used + recently traded ----------------------
    ctx5 = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx5)
    await journal.ask_market(upd.text("🪙 کریپتو"), ctx5)
    await journal.ask_symbol(upd.text("EURUSD"), ctx5)
    await journal.ask_direction(upd.text("l"), ctx5)
    state = await journal.ask_leverage(upd.text("⏭ بدون اهرم"), ctx5)
    check(
        state == journal.TIMEFRAME and "leverage" not in ctx5.user_data,
        "leverage skippable",
    )
    await journal.ask_timeframe(upd.text("5m"), ctx5)
    await journal.ask_entry(upd.text("1.1"), ctx5)
    await journal.ask_take_profit(upd.text("1.2"), ctx5)
    await journal.ask_stop_loss(upd.text("1.05"), ctx5)
    state = await journal.ask_result(upd.text("win"), ctx5)  # typed alias
    check(state == journal.MARGIN and ctx5.user_data["hit"] == "win", "typed 'win' alias")
    await journal.ask_margin(upd.text("1"), ctx5)
    state = await journal.ask_risk(upd.text("2"), ctx5)  # typed without %
    check(
        state == journal.TRADE_DATE and ctx5.user_data["risk_percent"] == 2,
        "typed risk without % accepted",
    )
    await journal.ask_trade_date(upd.text("-"), ctx5)
    state = await journal.ask_trade_hour(upd.text("🕐 الان"), ctx5)
    check(
        state == journal.NOTES
        and ctx5.user_data["trade_date"].startswith(date.today().isoformat())
        and " " in ctx5.user_data["trade_date"],
        "🕐 Now appends the current hour to the date",
    )
    await journal.ask_notes(upd.text("-"), ctx5)
    await journal.ask_mood(upd.text("⏭ رد کردن"), ctx5)
    await journal.ask_screenshot_text(upd.text("-"), ctx5)
    await journal.ask_screenshot_after_text(upd.text("-"), ctx5)
    state = await journal.save_trade(upd.text("y"), ctx5)
    check(state == journal.ConversationHandler.END, "third trade (EURUSD) saved")

    recent, top = db.get_symbol_suggestions()
    check(
        recent[0] == "EURUSD"
        and recent[1] == "BTCUSD"
        and top[0] == "EURUSD"
        and top[1] == "BTCUSD",
        "symbol suggestions: recent + most-used ordering",
    )

    ctx_kb = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx_kb)
    await journal.ask_market(upd.text("🪙 کریپتو"), ctx_kb)
    kb = upd.sent[-1][2]
    check(
        isinstance(kb, ReplyKeyboardMarkup),
        "symbol prompt shows a reply keyboard",
    )
    kb_labels = _labels(kb)
    check(
        "EURUSD" in kb_labels
        and "BTCUSD" in kb_labels
        and "✖️ لغو" in kb_labels,
        "most used + recently traded symbols appear as buttons",
    )
    check(
        kb_labels.index("EURUSD") < kb_labels.index("BTCUSD"),
        "most used symbol is listed first",
    )
    state = await journal.ask_symbol(upd.text("EURUSD"), ctx_kb)
    check(
        state == journal.DIRECTION and ctx_kb.user_data["symbol"] == "EURUSD",
        "tapping a symbol button proceeds to DIRECTION",
    )
    state = await journal.ask_symbol(upd.text("not a symbol"), ctx_kb)
    check(
        state == journal.SYMBOL
        and isinstance(upd.sent[-1][2], ReplyKeyboardMarkup),
        "invalid symbol re-prompts with the symbol buttons",
    )

    # --- discard removes the attached screenshot files ------------------------
    ctx3 = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx3)
    await journal.ask_market(upd.text("🪙 کریپتو"), ctx3)
    await journal.ask_symbol(upd.text("AAPL"), ctx3)
    await journal.ask_direction(upd.text("long"), ctx3)
    await journal.ask_leverage(upd.text("×5"), ctx3)
    await journal.ask_timeframe(upd.text("5m"), ctx3)
    await journal.ask_entry(upd.text("200"), ctx3)
    await journal.ask_take_profit(upd.text("210"), ctx3)
    await journal.ask_stop_loss(upd.text("190"), ctx3)
    await journal.ask_result(upd.text("❌ Loss"), ctx3)
    await journal.ask_margin(upd.text("1"), ctx3)
    await journal.ask_risk(upd.text("0.5%"), ctx3)
    await journal.ask_trade_date(upd.text("-"), ctx3)
    await journal.ask_trade_hour(upd.text("-"), ctx3)
    await journal.ask_notes(upd.text("-"), ctx3)
    await journal.ask_mood(upd.text("⏭ رد کردن"), ctx3)
    await journal.ask_screenshot(upd.photo(), ctx3)
    doomed = ctx3.user_data["screenshot"]
    check((journal.SCREENSHOT_DIR / doomed).is_file(), "before screenshot attached")
    await journal.ask_screenshot_after(upd.photo(), ctx3)
    doomed_after = ctx3.user_data["screenshot_after"]
    check(
        (journal.SCREENSHOT_DIR / doomed_after).is_file(),
        "after screenshot attached",
    )
    state = await journal.save_trade(upd.text("❌ ثبت نشود"), ctx3)
    check(
        state == journal.ConversationHandler.END and not ctx3.user_data,
        "Discard button -> END, draft cleared",
    )
    check(
        not (journal.SCREENSHOT_DIR / doomed).exists()
        and not (journal.SCREENSHOT_DIR / doomed_after).exists(),
        "discarded draft's screenshot files removed",
    )
    check(
        _labels(_last_markup(upd)) == _home_labels(),
        "discard restores the main-menu bar",
    )

    # --- ✖️ Cancel mid-flow also removes the screenshot ------------------------
    ctx4 = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx4)
    await journal.ask_market(upd.text("🪙 کریپتو"), ctx4)
    await journal.ask_symbol(upd.text("TSLA"), ctx4)
    await journal.ask_direction(upd.text("l"), ctx4)
    await journal.ask_leverage(upd.text("×10"), ctx4)
    await journal.ask_timeframe(upd.text("4h"), ctx4)
    await journal.ask_entry(upd.text("10"), ctx4)
    await journal.ask_take_profit(upd.text("11"), ctx4)
    await journal.ask_stop_loss(upd.text("9"), ctx4)
    state = await journal.ask_result(upd.text("✅ Win"), ctx4)
    check(
        state == journal.MARGIN and ctx4.user_data["exit_price"] == 11,
        "Win (ctx4)",
    )
    await journal.ask_margin(upd.text("3"), ctx4)
    await journal.ask_risk(upd.text("-"), ctx4)
    await journal.ask_trade_date(upd.text("-"), ctx4)
    await journal.ask_trade_hour(upd.text("-"), ctx4)
    await journal.ask_notes(upd.text("-"), ctx4)
    await journal.ask_mood(upd.text("⏭ رد کردن"), ctx4)
    await journal.ask_screenshot(upd.photo(), ctx4)
    doomed2 = ctx4.user_data["screenshot"]
    state = await journal.cancel(upd.text("✖️ لغو"), ctx4)
    check(state == journal.ConversationHandler.END, "✖️ Cancel -> END")
    check(
        not (journal.SCREENSHOT_DIR / doomed2).exists(),
        "cancelled draft's screenshot file removed",
    )

    # --- date given with time skips the hour question -------------------------
    ctx6 = FakeContext()
    await journal.trade_start(upd.text("/trade"), ctx6)
    await journal.ask_market(upd.text("🪙 کریپتو"), ctx6)
    await journal.ask_symbol(upd.text("MSFT"), ctx6)
    await journal.ask_direction(upd.text("short"), ctx6)
    await journal.ask_leverage(upd.text("×20"), ctx6)
    await journal.ask_timeframe(upd.text("1h"), ctx6)
    await journal.ask_entry(upd.text("400"), ctx6)
    await journal.ask_take_profit(upd.text("380"), ctx6)
    await journal.ask_stop_loss(upd.text("410"), ctx6)
    await journal.ask_result(upd.text("❌ Loss"), ctx6)
    await journal.ask_margin(upd.text("2"), ctx6)
    await journal.ask_risk(upd.text("-"), ctx6)
    state = await journal.ask_trade_date(upd.text("2026-02-01 09:30"), ctx6)
    check(
        state == journal.NOTES
        and ctx6.user_data["trade_date"] == "2026-02-01 09:30",
        "datetime at the date step skips the hour question",
    )
    await journal.ask_notes(upd.text("-"), ctx6)
    await journal.ask_mood(upd.text("آرام"), ctx6)
    await journal.ask_screenshot_text(upd.text("-"), ctx6)
    await journal.ask_screenshot_after_text(upd.text("-"), ctx6)
    state = await journal.save_trade(upd.text("y"), ctx6)
    check(state == journal.ConversationHandler.END, "fourth trade saved")
    msft_row = db.get_recent(1)[0]
    # Auto P&L for this short: 2 margin * 20x * ((400 - 410) / 400) = -1.0
    check(
        msft_row["symbol"] == "MSFT"
        and msft_row["trade_date"] == "2026-02-01 09:30"
        and msft_row["mood"] == "calm"
        and msft_row["hit"] == "lose"
        and msft_row["exit_price"] == 410
        and abs(msft_row["pnl"] - (-1.0)) < 1e-9
        and msft_row["market"] == "crypto"
        and msft_row["leverage"] == 20
        and msft_row["take_profit"] == 380
        and msft_row["stop_loss"] == 410
        and msft_row["size"] == 2,
        "datetime + mood stored in trade",
    )

    # --- 🗑 delete inside the recent detail view (disposable trade) -------------
    dummy_id = db.add_trade(
        symbol="SPY",
        direction="long",
        timeframe="1h",
        entry_price=10.0,
        exit_price=11.0,
        size=1.0,
        pnl=1.0,
        trade_date=date.today().isoformat(),
        notes="",
        mood="",
        roi=100.0,
        market="crypto",
        leverage=2,
        risk_percent=1,
        take_profit=11,
        stop_loss=9,
        hit="tp",
    )
    del_upd = await _tap_recent(f"rec:d:{dummy_id}")
    check(db.get_trade(dummy_id) is None, "🗑 tap inside detail deletes the trade")
    check(
        _reply_text(del_upd) == "🗑 معامله حذف شد ✅",
        "deleting a sent detail confirms on the detail message",
    )
    recheck = await _tap_recent("rec:r:all")
    check(
        f"#{dummy_id}" not in [
            t for t, _ in _inline_flat(_last_markup(recheck))
        ],
        "deleted trade no longer listed",
    )
    miss_upd = await _tap_recent("rec:d:424242")
    check(
        miss_upd.callback_query.answers == ["این معامله دیگر وجود ندارد."],
        "🗑 tap on a stale id explains",
    )

    # --- 🟢 panel: inline buttons, detail card, close flow ----------------------
    op_upd = FakeUpdate()
    await journal.open_trades(op_upd.text("/open"), FakeContext())
    op_markup = _last_markup(op_upd)
    check(
        isinstance(op_markup, InlineKeyboardMarkup)
        and op_upd.sent[-1][1].startswith("🟢"),
        "🟢 /open sends the paginated open-trades panel",
    )
    op_flat = _inline_flat(op_markup)
    check(
        (f"🟢 #{xau['id']} — XAUUSD · 2000 📷", f"opn:v:{xau['id']}") in op_flat
        and "opn:home" in [cb for _, cb in op_flat],
        "open panel: one button per open trade with entry price + pager",
    )
    check(
        ("➕ ثبت معامله باز", "opn:add") in op_flat,
        "open panel carries the ➕ add button",
    )

    class FakeOpenQuery:
        def __init__(self, data, upd):
            self.data = data
            self._upd = upd
            self.answers = []

        async def answer(self, text=None):
            self.answers.append(text)

        async def edit_message_text(self, text, reply_markup=None, **kwargs):
            self._upd.sent.append(("edit", text, reply_markup))

        async def edit_message_reply_markup(self, reply_markup=None, **kwargs):
            self._upd.sent.append(("edit-markup", "", reply_markup))

    class FakeOpenCallbackUpdate:
        def __init__(self, data):
            self.sent = []
            self.effective_chat = FakeChat(self.sent)
            self.message = None
            self.callback_query = FakeOpenQuery(data, self)

    async def _tap_open(data):
        cb_upd = FakeOpenCallbackUpdate(data)
        await journal.on_open_callback(cb_upd, FakeContext())
        return cb_upd

    odet = await _tap_open(f"opn:v:{xau['id']}")
    kind_, text_, markup_ = odet.sent[-1]
    check(
        kind_ == "send" and f"معامله باز #{xau['id']} — <b>XAUUSD</b>" in text_,
        "tapping an open trade SENDS its detail card",
    )
    check(
        "Entry: <code>2000</code>" in text_
        and "🎯 TP: <code>2100</code>" in text_
        and "🛑 SL: <code>1950</code>" in text_
        and "Risk: 1%" in text_
        and "2026-03-01 10:30" in text_
        and "broke range high" in text_,
        "open detail card shows entry/TP/SL/risk/date/reason",
    )
    check(
        ("🏁 Close trade", f"opn:c:{xau['id']}") in _inline_flat(markup_)
        and ("🗑 حذف", f"opn:d:{xau['id']}") in _inline_flat(markup_)
        and ("❌ بستن", "opn:close") in _inline_flat(markup_)
        and ("📷 عکس چارت", f"opn:ph:{xau['id']}") in _inline_flat(markup_),
        "open detail card has 🏁 Close, 🗑 delete, ❌ close and 📷 buttons",
    )
    oph = await _tap_open(f"opn:ph:{xau['id']}")
    check(
        any(s[0] == "photo" and "چارت ورود" in s[1] for s in oph.sent),
        "📷 button on the open detail sends the entry screenshot",
    )

    # --- 🏁 close flow: TP-hit path with two exit photos -------------------------
    cctx = FakeContext()
    state = await journal._close_begin(
        xau["id"], upd.text(f"/close {xau['id']}"), cctx
    )
    check(
        state == journal.CLOSE_STATUS
        and cctx.user_data["open_id"] == xau["id"]
        and cctx.user_data["open_symbol"] == "XAUUSD",
        "/close <id> starts CLOSE_STATUS",
    )
    state = await journal.ask_close_status(upd.text("✅ Win (TP)"), cctx)
    check(
        state == journal.CLOSE_DATE and cctx.user_data["hit"] == "win",
        "close: Win (TP) -> CLOSE_DATE",
    )
    state = await journal.ask_close_date(upd.text("2026-03-03"), cctx)
    check(state == journal.CLOSE_HOUR, "close: exit date -> CLOSE_HOUR (separate)")
    state = await journal.ask_close_hour(upd.text("15:45"), cctx)
    check(
        state == journal.CLOSE_PHOTOS and cctx.user_data["exit_price"] == 2100,
        "close: TP-hit auto-fills exit price 2100, -> CLOSE_PHOTOS",
    )

    state = await journal.ask_close_photos(upd.photo(), cctx)
    check(
        state == journal.CLOSE_PHOTOS
        and len(cctx.user_data["exit_photos"].splitlines()) == 1,
        "close: 1st exit photo stored (asks for more)",
    )
    state = await journal.ask_close_photos(upd.photo(), cctx)
    check(
        state == journal.CLOSE_PHOTOS
        and len(cctx.user_data["exit_photos"].splitlines()) == 2,
        "close: 2nd exit photo stored",
    )
    state = await journal.ask_close_photos_text(upd.text("⏭ بدون اسکرین‌شات"), cctx)
    check(state == journal.CLOSE_REASON, "close: skip photos -> CLOSE_REASON")
    state = await journal.ask_close_reason(
        upd.text("TP tapped, momentum gone"), cctx
    )
    check(
        state == journal.CLOSE_MOOD
        and cctx.user_data["notes"] == "TP tapped, momentum gone",
        "close: exit reason -> CLOSE_MOOD (close-specific prompt, not /trade MOOD)",
    )
    state = await journal.ask_close_mood(upd.text("آرام"), cctx)
    check(
        state == journal.CLOSE_CONFIRM and cctx.user_data["mood"] == "calm",
        "close: mood -> CLOSE_CONFIRM",
    )
    csummary = upd.sent[-1][1]
    check(
        "Status" in csummary and "TP hit (Win)" in csummary
        and "Exit" in csummary and "2100" in csummary
        and "Date" in csummary and "2026-03-03 15:45" in csummary
        and "Shots" in csummary and "۲" in csummary
        and "Reason" in csummary and "Mood" in csummary and "آرام" in csummary,
        "close: confirmation shows status/exit/date/shots/reason/mood",
    )
    state = await journal.ask_close_mood(upd.text("خورشت قیمه"), cctx)
    check(
        state == journal.CLOSE_MOOD,
        "close: invalid mood re-asks CLOSE_MOOD (stays routable)",
    )
    state = await journal.save_close_trade(upd.text("✅ ثبت"), cctx)
    check(
        state == journal.ConversationHandler.END and not cctx.user_data,
        "close: save -> END, draft cleared",
    )
    check(
        db.count_open_trades() == 0 and db.get_open_trade(xau["id"]) is None,
        "closed trade removed from open_trades",
    )
    closed = db.get_recent(1)[0]
    check(
        closed["symbol"] == "XAUUSD" and closed["source"] == "open"
        and closed["hit"] == "win" and closed["exit_price"] == 2100
        and closed["entry_price"] == 2000
        and closed["trade_date"] == "2026-03-03"
        and closed["exit_time"] == "15:45"
        and closed["entry_time"] == "10:30"
        and closed["entry_reason"] == "broke range high"
        and closed["notes"] == "TP tapped, momentum gone"
        and closed["exit_reason"] == "TP tapped, momentum gone"
        and closed["mood"] == "calm"
        and closed["screenshot"] == open_shot
        and len((closed["exit_photos"] or "").splitlines()) == 2
        and closed["pnl"] is None and closed["size"] is None
        and closed["roi"] is None,
        "closed row keeps both reasons/times/photos and NULL P&L",
    )
    st_after = db.get_stats()
    check(
        st_after["wins"] == 3 and st_after["losses"] == 2 and st_after["be"] == 0,
        "open-flow closes count as win/loss/BE (NULL P&L classified by hit)",
    )

    # --- 🏁 close flow: manual exit via /close + skip tokens ----------------------
    oid2 = db.add_open_trade(
        symbol="BTCUSD", direction="short", market="crypto", timeframe="4h",
        reason="", screenshot=None, trade_date="2026-03-02", entry_time="08:00",
        risk_percent=2, entry_price=60000, take_profit=57000, stop_loss=61500,
    )
    cctx2 = FakeContext()
    state = await journal.close_start_text(upd.text("/close"), cctx2)
    check(
        state == journal.ConversationHandler.END,
        "/close without an id explains usage",
    )
    cctx2 = FakeContext(args=[str(oid2)])
    state = await journal.close_start_text(upd.text("/close"), cctx2)
    state = await journal.close_start_text(
        upd.text(f"/close {oid2}"), cctx2
    )
    check(
        state == journal.CLOSE_STATUS and cctx2.user_data["open_id"] == oid2,
        "/close <id> enters the close flow",
    )
    state = await journal.ask_close_status(upd.text("✏️ Manual"), cctx2)
    check(
        state == journal.CLOSE_DATE and cctx2.user_data["hit"] == "manual",
        "close: Manual -> CLOSE_DATE",
    )
    await journal.ask_close_date(upd.text("2026-03-04"), cctx2)
    state = await journal.ask_close_hour(upd.text("-"), cctx2)
    check(state == journal.CLOSE_PRICE, "close: manual exit asks for the price")
    state = await journal.ask_close_price(upd.text("58500"), cctx2)
    check(
        state == journal.CLOSE_PHOTOS
        and cctx2.user_data["exit_price"] == 58500,
        "close: manual price stored",
    )
    state = await journal.ask_close_photos_text(upd.text("-"), cctx2)
    check(state == journal.CLOSE_REASON, "close: no exit photos -> reason")
    await journal.ask_close_reason(upd.text("-"), cctx2)
    state = await journal.ask_close_mood(upd.text("⏭ رد کردن"), cctx2)
    check(state == journal.CLOSE_CONFIRM, "close: mood skip -> confirm")
    state = await journal.save_close_trade(upd.text("y"), cctx2)
    check(state == journal.ConversationHandler.END, "close: manual trade saved")
    manual = db.get_recent(1)[0]
    check(
        manual["hit"] == "manual" and manual["exit_price"] == 58500
        and manual["exit_time"] == "" and manual["trade_date"] == "2026-03-04",
        "manual exit stored with typed price and skipped time",
    )

    # --- 🏁 close flow: BE auto-fills the entry price ------------------------------
    oid3 = db.add_open_trade(
        symbol="ETHUSD", direction="long", market="crypto", timeframe="1h",
        reason="", screenshot=None, trade_date="2026-03-02", entry_time="",
        risk_percent=None, entry_price=3000, take_profit=3200, stop_loss=2900,
    )
    cctx3 = FakeContext()
    await journal._close_begin(oid3, upd.text("x"), cctx3)
    await journal.ask_close_status(upd.text("➖ BE"), cctx3)
    await journal.ask_close_date(upd.text("2026-03-05"), cctx3)
    state = await journal.ask_close_hour(upd.text("12:00"), cctx3)
    check(
        state == journal.CLOSE_PHOTOS
        and cctx3.user_data["exit_price"] == 3000,
        "close: BE auto-fills the entry price as exit",
    )
    await journal.ask_close_photos_text(upd.text("-"), cctx3)
    await journal.ask_close_reason(upd.text("-"), cctx3)
    state = await journal.ask_close_mood(upd.text("-"), cctx3)
    check(state == journal.CLOSE_CONFIRM, "close: BE mood skipped -> confirm")
    state = await journal.save_close_trade(upd.text("✅ ثبت"), cctx3)
    check(state == journal.ConversationHandler.END, "close: BE trade saved")
    be_row = db.get_recent(1)[0]
    check(
        be_row["hit"] == "be" and be_row["exit_price"] == 3000
        and be_row["risk_percent"] is None,
        "BE close stored with the entry price as exit",
    )

    # --- 🟢 /open with nothing open: empty panel with ➕ (no auto-start) ---------
    empty_upd = FakeUpdate()
    await journal.open_trades(empty_upd.text("/open"), FakeContext())
    empty_flat = _inline_flat(_last_markup(empty_upd))
    check(
        ("➕ ثبت معامله باز", "opn:add") in empty_flat
        and "باز نیست" in empty_upd.sent[-1][1],
        "empty 🟢 panel shows ➕ (no synthetic questionnaire start)",
    )

    # --- mood parsing details ---------------------------------------------------
    ctx7 = FakeContext()
    await journal._prompt_mood(upd)
    state = await journal.ask_mood(upd.text("banana"), ctx7)
    check(state == journal.MOOD, "invalid mood reprompts")
    state = await journal.ask_mood(upd.text("fomo"), ctx7)
    check(
        state == journal.SCREENSHOT and ctx7.user_data["mood"] == "fomo",
        "typed mood alias accepted",
    )

    # --- mood parsing details ---------------------------------------------------
    check(
        [s for s, _ in db.get_all_symbols()][:3]
        == ["ETHUSD", "BTCUSD", "XAUUSD"],
        "symbols sorted by last trade (picker order, open-flow closes included)",
    )
    breakdown = db.get_mood_breakdown()
    check(
        [(r["mood"], r["trades"], r["total"]) for r in breakdown]
        == [("anxious", 1, 2.0), ("calm", 2, -1.0)],
        "mood breakdown ordered by P&L",
    )
    stats_all = db.get_stats()
    check(
        stats_all["trades"] == 7 and stats_all["be"] == 1,
        "stats totals + BE count",
    )
    check(db.get_stats(symbol="EURUSD")["trades"] == 2, "stats filter by symbol")
    check(db.get_stats(since="2999-01-01")["trades"] == 0, "stats filter by period")

    def _cb_update(data, message_id=400):
        """Fake callback-query update for the stats inline buttons."""
        query = SimpleNamespace(
            data=data,
            answers=[],
            message=SimpleNamespace(message_id=message_id),
        )

        async def _answer(text=None, show_alert=False):
            query.answers.append(text)

        async def _delete():
            query.deleted = True

        async def _edit(text, reply_markup=None, parse_mode=None, **kw):
            stats_upd.sent.append(("cb-edit", text, reply_markup))

        query.answer = _answer
        query.message.delete = _delete
        query.edit_message_text = _edit
        return SimpleNamespace(
            callback_query=query, effective_chat=stats_upd.effective_chat
        )

    stats_upd = FakeUpdate()
    stats_ctx = FakeContext(["BTCUSD", "3m"])
    stats_ctx.bot = FakeBot(stats_upd.sent)
    await journal.stats(stats_upd.text("/stats BTCUSD 3m"), stats_ctx)
    panel = stats_upd.sent[-1][1]
    check(
        "BTCUSD" in panel and "سه ماه گذشته" in panel and "Trades:" in panel,
        "stats header + filtered panel text",
    )
    check(
        "Avg ROI:" in panel and "-20.00%" in panel,
        "stats panel shows the average ROI",
    )
    check(
        "◾" in panel and panel.count("•") >= 5,
        "stats panel uses section markers and bullet rows",
    )
    kb = stats_upd.sent[-1][2]
    check(
        isinstance(kb, InlineKeyboardMarkup),
        "stats filters are inline (attached to the message, not the reply bar)",
    )
    flat = _inline_flat(kb)
    check(
        ("✓ 3M", "stat:p:3m") in flat
        and ("1M", "stat:p:1m") in flat
        and ("All", "stat:p:all") in flat
        and ("🔤 Symbols: BTCUSD", "stat:open") in flat
        and ("♻️ Reset", "stat:reset") in flat
        and ("📤 Export", "stat:export") in flat,
        "stats filter buttons present (periods, symbols, reset, export)",
    )

    # period tap -> the SAME message is edited in place
    await journal.on_stats_callback(_cb_update("stat:p:1m"), stats_ctx)
    check(
        stats_ctx.user_data["stats_filter"]["period"] == "1m"
        and stats_ctx.user_data["stats_filter"]["symbol"] == "BTCUSD",
        "period tap updates the filter, keeps the symbol",
    )
    last = stats_upd.sent[-1]
    check(
        last[0] == "cb-edit" and "سی روز گذشته" in last[1],
        "period tap re-renders the panel in place (no new messages)",
    )

    # Symbols tap -> the picker is its OWN message with paged symbol buttons
    await journal.on_stats_callback(_cb_update("stat:open"), stats_ctx)
    last = stats_upd.sent[-1]
    check(
        last[0] == "bot-send" and "نمادها" in last[1],
        "Symbols tap sends a separate symbol-picker message",
    )
    pk = last[2]
    pk_labels = [b.text for row in pk.inline_keyboard for b in row]
    check(
        pk_labels[0] == "ETHUSD (1)"
        and any(lbl.startswith("BTCUSD") for lbl in pk_labels),
        "picker lists symbols sorted by last trade with counts",
    )
    await journal.on_stats_callback(_cb_update("stat:syms:1"), stats_ctx)
    check(
        stats_upd.sent[-1][0] == "cb-edit",
        "page navigation edits the picker in place",
    )

    # symbol tap -> filter set, picker closed, panel refreshed in place
    await journal.on_stats_callback(_cb_update("stat:reset"), stats_ctx)
    await journal.on_stats_callback(_cb_update("stat:sym:BTCUSD"), stats_ctx)
    check(
        stats_ctx.user_data["stats_filter"]["symbol"] == "BTCUSD"
        and stats_upd.sent[-1][0] == "bot-edit"
        and "BTCUSD" in stats_upd.sent[-1][1],
        "symbol tap filters and refreshes the panel in place",
    )
    await journal.on_stats_callback(_cb_update("stat:sym:BTCUSD"), stats_ctx)
    check(
        stats_ctx.user_data["stats_filter"]["symbol"] is None,
        "tapping the active symbol again clears it",
    )

    # reset clears everything
    await journal.on_stats_callback(_cb_update("stat:p:1m"), stats_ctx)
    await journal.on_stats_callback(_cb_update("stat:reset"), stats_ctx)
    check(
        stats_ctx.user_data["stats_filter"] == {"symbol": None, "period": None},
        "Reset clears all filters",
    )

    # picker pagination math: exactly 10 symbols per page
    fake_syms = [(f"SYM{i:02d}", i) for i in range(1, 26)]
    kb2 = journal._symbol_picker_kb(2, 3, fake_syms)
    labels2 = [b.text for row in kb2.inline_keyboard for b in row]
    check(
        "SYM11 (11)" in labels2
        and "SYM20 (20)" in labels2
        and "SYM10 (10)" not in labels2
        and "SYM21 (21)" not in labels2,
        "picker shows exactly 10 symbols per page",
    )
    cbs2 = [b.callback_data for row in kb2.inline_keyboard for b in row]
    check(
        "stat:syms:1" in cbs2 and "stat:syms:3" in cbs2,
        "picker page navigation buttons",
    )

    # typed /stats arguments still parse
    stats_ctx.args = ["eurusd", "1w"]
    await journal.stats(stats_upd.text("/stats eurusd 1w"), stats_ctx)
    check(
        "EURUSD" in stats_upd.sent[-1][1]
        and "هفت روز گذشته" in stats_upd.sent[-1][1]
        and stats_ctx.user_data["stats_filter"]["symbol"] == "EURUSD"
        and stats_ctx.user_data["stats_filter"]["period"] == "1w",
        "/stats eurusd 1w parses symbol + period",
    )

    # export via the panel's inline button
    await journal.on_stats_callback(_cb_update("stat:export"), stats_ctx)
    check(
        stats_ctx.bot.sent_docs
        and stats_ctx.bot.sent_docs[0][0].startswith("trades-")
        and stats_ctx.bot.sent_docs[0][2][:2] == b"PK",
        "Export button sends the .xlsx document",
    )

    # --- spreadsheet export ------------------------------------------------------
    export_path = export.build_export_file()
    check(
        export_path.parent == export.EXPORT_DIR
        and export_path.name.startswith("trades-")
        and export_path.suffix == ".xlsx",
        "export file created with timestamped name",
    )
    from openpyxl import load_workbook

    wb = load_workbook(export_path)
    ws = wb.active
    check(
        [c.value for c in ws[1]][:5]
        == ["ID", "Symbol", "Market", "Direction", "Timeframe"],
        "export header row present",
    )
    check(ws.max_row == 8, "export has one row per trade (7 trades)")
    check(
        "Screenshot" not in [c.value for c in ws[1]] and ws.max_column == 19,
        "export has no screenshot columns (photos stay in Telegram)",
    )
    row2 = [c.value for c in ws[2]]
    check(
        row2[0] == 1
        and row2[1] == "EURUSD"
        and row2[2] == "crypto"
        and row2[3] == "LONG"
        and row2[4] == "1h"
        and row2[7] == 110  # take profit
        and row2[8] == 95  # stop loss
        and row2[9] == "win"  # result (stored in the hit column)
        and row2[10] == 10  # leverage
        and row2[11] == 2  # margin
        and row2[12] == 1  # risk %
        and row2[13] == rows[1]["pnl"]  # P&L
        and row2[14] == rows[1]["roi"] == 100.0  # ROI %
        and row2[15] == rows[1]["trade_date"]  # date moved after ROI
        and row2[16] == "anxious",
        "export row values match the stored trade",
    )
    check(ws["N2"].number_format == "0.####", "P&L column stays numeric")
    check(
        ws.freeze_panes == "A2" and bool(ws.auto_filter.ref),
        "header frozen and filterable",
    )
    wb.close()
    export_path.unlink(missing_ok=True)  # keep the dir empty for the flow test

    class FakeDocumentReplyMessage(FakeMessage):
        def __init__(self, log):
            super().__init__(text=None, log=log)
            self.documents = []

        async def reply_document(
            self, document, filename=None, caption=None, reply_markup=None
        ):
            self.documents.append((filename, caption, reply_markup))
            self.last_document_bytes = document.read()

        async def reply_chat_action(self, action):
            self.chat_actions = getattr(self, "chat_actions", []) + [action]

    doc_upd = FakeUpdate()
    doc_upd.message = FakeDocumentReplyMessage(doc_upd.sent)

    await journal.export_trades(doc_upd, FakeContext())
    filename, caption, markup = doc_upd.message.documents[0]
    check(
        filename.startswith("trades-") and filename.endswith(".xlsx"),
        "/export sends a timestamped .xlsx document",
    )
    check(
        doc_upd.message.last_document_bytes[:2] == b"PK",
        "export bytes are a real xlsx (zip) file",
    )
    check("همه معاملات" in caption, "export caption present")
    check(
        markup is journal._MENU_KEYBOARD,
        "export document re-attaches the main-menu bar",
    )
    check(
        list(export.EXPORT_DIR.glob("trades-*.xlsx")) == [],
        "export temp file removed after sending",
    )

    # --- 📷 chart-photo button on the /recent detail card -----------------------
    class FakePhotoReplyMessage(FakeMessage):
        def __init__(self, log):
            super().__init__(text=None, log=log)
            self.photos = []

        async def reply_photo(self, photo, caption=None):
            self.photos.append((photo, caption))

    all_rows = db.get_recent(10)
    shot_row = next(r for r in all_rows if r["screenshot"])
    noshot_row = next(r for r in all_rows if not r["screenshot"])
    check(
        journal._RECENT_CB_RE.match(f"rec:ph:{shot_row['id']}") is not None,
        "rec:ph callback tokens are routed by the recent-panel regex",
    )
    # Only trades WITH screenshots get the 📷 button on their detail card.
    check(
        all(
            not cb.startswith("rec:ph:")
            for _, cb in _inline_flat(journal._recent_detail_kb(noshot_row["id"]))
        ),
        "detail of a shotless trade has no 📷 button",
    )
    check(
        ("📷 عکس چارت", f"rec:ph:{shot_row['id']}")
        in _inline_flat(journal._recent_detail_kb(shot_row["id"], True)),
        "detail of a trade with screenshots carries the 📷 button",
    )
    # Tapping 📷 SENDS the stored screenshots as photos.
    ph_upd = await _tap_recent(f"rec:ph:{shot_row['id']}")
    photo_sends = [s for s in ph_upd.sent if s[0] == "photo"]
    check(
        len(photo_sends) == 1
        and photo_sends[0][1].startswith(f"#{shot_row['id']}")
        and "قبل از معامله" in photo_sends[0][1],
        "📷 button sends the stored screenshot (قبل از معامله)",
    )
    # A trade without shots explains instead of sending photos.
    ph_upd2 = await _tap_recent(f"rec:ph:{noshot_row['id']}")
    check(
        any("اسکرین‌شات ندارد" in s[1] for s in ph_upd2.sent if s[0] == "send"),
        "📷 button on a shotless trade explains",
    )

    # --- pagination: 11 trades -> two pages + range filter ----------------------
    for i in range(7):
        db.add_trade(
            symbol=f"PG{i}",
            direction="long",
            timeframe="1h",
            entry_price=1.0,
            exit_price=1.1,
            size=1.0,
            pnl=0.1,
            trade_date=date.today().isoformat(),
            notes="",
            mood="",
            roi=10.0,
            market="crypto",
            leverage=1,
            hit="tp",
        )
    pg_upd = FakeUpdate()
    await journal.recent(pg_upd.text("/recent"), FakeContext())
    pg_cbs = [cb for _, cb in _inline_flat(_last_markup(pg_upd))]
    check(
        "rec:p:2" in pg_cbs,
        "with 11 trades the panel offers page 2",
    )
    pg2 = await _tap_recent("rec:p:2")
    pg2_markup = _last_markup(pg2)
    pg2_labels = [t for t, _ in _inline_flat(pg2_markup)]
    check(
        f"{journal._fa_num(2)} / {journal._fa_num(2)}" in pg2_labels,
        "page indicator shows ۲ / ۲",
    )
    pg2_cbs = [cb for _, cb in _inline_flat(pg2_markup)]
    check(
        "rec:p:1" in pg2_cbs and "rec:p:2" not in pg2_cbs,
        "page 2 has back but no forward/self page button",
    )
    rng = await _tap_recent("rec:r:1w")
    check(
        "✓ 1W" in [t for t, _ in _inline_flat(_last_markup(rng))],
        "range tap marks the button active and re-renders the panel",
    )

    # --- conversation wiring (real PTB routing, no network) -------------------
    conv = journal.build_conversation()
    _user = User(id=1, first_name="T", is_bot=False)
    _chat = Chat(id=1, type=Chat.PRIVATE)

    class _BotStub:
        """Just enough bot for CommandHandler's username check."""

        username = "TestBot"

    def _text_update(text):
        kwargs = {}
        if text.startswith("/"):
            # real clients always tag "/cmd" texts with a bot_command entity,
            # and PTB's CommandHandler asks message.get_bot() for the bot's
            # username — provide both like a real update would have.
            kwargs["entities"] = [
                MessageEntity(
                    type=MessageEntity.BOT_COMMAND,
                    offset=0,
                    length=len(text.split()[0]),
                )
            ]
        msg = Message(
            message_id=2,
            date=_dt.now(),
            chat=_chat,
            from_user=_user,
            text=text,
            **kwargs,
        )
        if text.startswith("/"):
            msg.set_bot(_BotStub())
        return Update(update_id=2, message=msg)

    def _photo_update():
        msg = Message(
            message_id=3,
            date=_dt.now(),
            chat=_chat,
            from_user=_user,
            photo=[object()],
        )
        return Update(update_id=3, message=msg)

    key = conv._get_key(_text_update("long"))

    def _routed_to(st, update):
        conv._conversations[key] = st
        result = conv.check_update(update)
        conv._conversations.clear()
        return result[2] if result else None

    cases = [
        (journal.MARKET, "🪙 کریپتو", journal.ask_market),
        (journal.SYMBOL, "EURUSD", journal.ask_symbol),
        (journal.DIRECTION, "📈 خرید", journal.ask_direction),
        (journal.LEVERAGE, "×10", journal.ask_leverage),
        (journal.TIMEFRAME, "5m", journal.ask_timeframe),
        (journal.ENTRY, "100", journal.ask_entry),
        (journal.TAKE_PROFIT, "110", journal.ask_take_profit),
        (journal.STOP_LOSS, "95", journal.ask_stop_loss),
        (journal.RESULT, "✅ Win", journal.ask_result),
        (journal.MARGIN, "2", journal.ask_margin),
        (journal.RISK, "1%", journal.ask_risk),
        (journal.TRADE_DATE, "2026-02-09", journal.ask_trade_date),
        (journal.TRADE_HOUR, "14:30", journal.ask_trade_hour),
        (journal.NOTES, "broke the low", journal.ask_notes),
        (journal.MOOD, "فومو", journal.ask_mood),
        (journal.SCREENSHOT, "-", journal.ask_screenshot_text),
        (journal.SCREENSHOT_AFTER, "-", journal.ask_screenshot_after_text),
        (journal.CONFIRM, "✅ ذخیره", journal.save_trade),
    ]
    for st, text, fn in cases:
        handler = _routed_to(st, _text_update(text))
        check(
            getattr(handler, "callback", None) is fn,
            f"PTB routes '{text}' at state {st} to {fn.__name__}",
        )

    handler = _routed_to(journal.SCREENSHOT, _photo_update())
    check(
        getattr(handler, "callback", None) is journal.ask_screenshot,
        "PTB routes a photo at SCREENSHOT to ask_screenshot",
    )

    handler = _routed_to(journal.ENTRY, _text_update("cancel"))
    check(
        getattr(handler, "callback", None) is journal.cancel,
        "typed 'cancel' mid-flow routes to cancel",
    )

    handler = _routed_to(journal.CONFIRM, _text_update("✖️ لغو"))
    check(
        getattr(handler, "callback", None) is journal.cancel,
        "✖️ لغو button routes to cancel",
    )

    def _entry_handler(update):
        result = conv.check_update(update)
        conv._conversations.clear()
        return result[2] if result else None

    handler = _entry_handler(_text_update("📈 معامله جدید"))
    check(
        getattr(handler, "callback", None) is journal.trade_start,
        "📈 معامله جدید tap enters the conversation (registers SYMBOL)",
    )
    handler = _entry_handler(_text_update("/trade"))
    check(
        getattr(handler, "callback", None) is journal.trade_start,
        "/trade command still enters the conversation",
    )
    handler = _routed_to(journal.DIRECTION, _text_update("📈 معامله جدید"))
    check(
        getattr(handler, "callback", None) is journal.trade_start,
        "📈 معامله جدید mid-conversation re-enters and restarts",
    )

    # --- ☰ menu-button command list + main menu -------------------------------
    menu_upd = FakeUpdate()
    menu_upd.text("/start")
    await journal.show_menu(menu_upd, FakeContext())
    menu_text = menu_upd.sent[-1][1]
    check(
        all(
            cmd in menu_text
            for cmd in (
                "/trade", "/open", "/opens", "/recent",
                "/stats", "/delete", "/cancel",
            )
        ),
        "menu lists every command with an explanation",
    )
    menu_kb = menu_upd.sent[-1][2]
    check(
        isinstance(menu_kb, ReplyKeyboardMarkup)
        and bool(menu_kb.is_persistent)
        and not menu_kb.one_time_keyboard,
        "main-menu keyboard is persistent (stays after taps)",
    )
    menu_labels = _labels(menu_kb)
    check(
        menu_labels == _home_labels(),
        "main-menu buttons present",
    )

    menu_handlers = journal.build_menu_handlers()

    def _menu_routes(label):
        """Callbacks of every menu handler matching label (truthy check!)."""
        return [
            h.callback for h in menu_handlers if h.check_update(_text_update(label))
        ]

    expected_menu = {
        "📊 آمار": journal.stats,
        "📊 Stats": journal.stats,  # English aliases still route
        "🟢 معاملات باز": journal.open_trades,
        "🟢 Open trades": journal.open_trades,
        "🕘 معاملات اخیر": journal.recent,
        "🕘 Recent": journal.recent,
        "📥 اکسل": journal.export_trades,
        "📥 Export": journal.export_trades,
        "🏠 منو": journal.show_menu,
        "🏠 Menu": journal.show_menu,
        "❓ راهنما": journal.show_menu,
        "❓ help": journal.show_menu,
    }
    for label, fn in expected_menu.items():
        check(
            _menu_routes(label) == [fn],
            f"menu button '{label}' routes to {fn.__name__}",
        )
    check(
        _menu_routes("📈 بستن معامله") == [],
        "📈 بستن معامله is handled by the conversation, not a standalone handler",
    )
    check(
        _menu_routes("🟢 ثبت معامله باز") == [],
        "🟢 ثبت معامله باز is handled by the conversation, not a standalone handler",
    )

    for label in ("hello", "new trade", "stats", "📈", "📊 stats now", "✖️", "📥"):
        check(
            _menu_routes(label) == [],
            f"non-menu text '{label}' is not routed by the menu handlers",
        )

    # --- stats inline callbacks: routing + real PTB handler ---------------------
    stats_handler = journal.build_stats_callbacks()
    check(
        isinstance(stats_handler, CallbackQueryHandler),
        "stats buttons are wired through a CallbackQueryHandler (inline)",
    )
    for data in (
        "stat:p:1w", "stat:p:all", "stat:sym:BTCUSD", "stat:syms:2",
        "stat:open", "stat:sall", "stat:reset", "stat:export",
        "stat:close", "stat:noop",
    ):
        check(
            bool(journal._STATS_CB_RE.match(data)),
            f"stats callback data '{data}' routes",
        )
    for data in (
        "hello", "stats", "stat:", "stat:sym:bad!", "statx:p:1w",
        "stat:p:", "BTCUSD (1)", "stats:p:1w",
    ):
        check(
            not journal._STATS_CB_RE.match(data),
            f"non-stats callback data '{data}' is ignored",
        )

    # --- recent inline callbacks: routing + real PTB handler -------------------
    recent_handler = journal.build_recent_callbacks()
    check(
        isinstance(recent_handler, CallbackQueryHandler),
        "recent buttons are wired through a CallbackQueryHandler (inline)",
    )
    for data in (
        "rec:p:1", "rec:r:1w", "rec:r:1m", "rec:r:all", "rec:v:3",
        "rec:ph:3", "rec:d:3", "rec:home", "rec:close", "rec:noop",
    ):
        check(
            bool(journal._RECENT_CB_RE.match(data)),
            f"recent callback data '{data}' routes",
        )
    for data in (
        "hello", "recent", "rec:", "rec:p:", "recx:p:1", "rec:r:1y",
        "rec:v:x", "recent:p:1",
    ):
        check(
            not journal._RECENT_CB_RE.match(data),
            f"non-recent callback data '{data}' is ignored",
        )

    # --- 🟢 open/close conversation wiring (real PTB routing) -------------------
    def _cb_update(data):
        """A real CallbackQuery update, like an inline-button tap produces."""
        msg = Message(message_id=9, date=_dt.now(), chat=_chat, from_user=_user)
        cq = CallbackQuery(
            id="cb1",
            from_user=_user,
            chat_instance="ci",
            data=data,
            message=msg,
        )
        return Update(update_id=9, callback_query=cq)

    open_conv = journal.build_open_conversation()

    def _open_entry(update):
        result = open_conv.check_update(update)
        open_conv._conversations.clear()
        return result[2] if result else None

    handler = _open_entry(_cb_update("opn:add"))
    check(
        getattr(handler, "callback", None) is journal.open_trades_add_entry,
        "➕ panel button enters the open conversation (entry point)",
    )
    handler = _open_entry(_text_update("🟢 ثبت معامله باز"))
    check(
        getattr(handler, "callback", None) is journal.open_trade_start,
        "🟢 ثبت معامله باز menu button starts the open questionnaire directly",
    )
    handler = _open_entry(_text_update("/open"))
    check(
        getattr(handler, "callback", None) is journal.open_trade_start,
        "/open starts the open questionnaire (/opens is the panel)",
    )
    check(
        _open_entry(_cb_update("opn:noop")) is None,
        "other opn taps don't start the conversation",
    )
    check(
        _open_entry(_text_update("➕ ثبت معامله باز")) is None,
        "plain ➕ text no longer starts the flow (button-only entry)",
    )
    check(
        _open_entry(_text_update("🟢 معاملات باز")) is None,
        "🟢 menu label does not enter the conversation (opens the panel)",
    )

    okey = open_conv._get_key(_text_update("long"))

    def _routed_open(st, update):
        open_conv._conversations[okey] = st
        result = open_conv.check_update(update)
        open_conv._conversations.clear()
        return result[2] if result else None

    open_cases = [
        (journal.OPEN_MARKET, "🪙 کریپتو", journal.ask_open_market),
        (journal.OPEN_SYMBOL, "XAUUSD", journal.ask_open_symbol),
        (journal.OPEN_DIRECTION, "📈 Long", journal.ask_open_direction),
        (journal.OPEN_TIMEFRAME, "1h", journal.ask_open_timeframe),
        (journal.OPEN_REASON, "broke out", journal.ask_open_reason),
        (journal.OPEN_SCREENSHOT, "-", journal.ask_open_screenshot_text),
        (journal.OPEN_TRADE_DATE, "2026-03-01", journal.ask_open_trade_date),
        (journal.OPEN_TRADE_HOUR, "10:30", journal.ask_open_trade_hour),
        (journal.OPEN_TRADE_HOUR, "الان", journal.ask_open_trade_hour),
        (journal.OPEN_RISK, "1%", journal.ask_open_risk),
        (journal.OPEN_ENTRY, "2000", journal.ask_open_entry),
        (journal.OPEN_TAKE_PROFIT, "2100", journal.ask_open_take_profit),
        (journal.OPEN_STOP_LOSS, "1950", journal.ask_open_stop_loss),
        (journal.OPEN_CONFIRM, "✅ ثبت", journal.save_open_trade),
    ]
    for st, text, fn in open_cases:
        handler = _routed_open(st, _text_update(text))
        check(
            getattr(handler, "callback", None) is fn,
            f"open conv routes '{text}' to {fn.__name__}",
        )

    close_conv = journal.build_close_conversation()

    def _close_entry(update):
        result = close_conv.check_update(update)
        close_conv._conversations.clear()
        return result[2] if result else None

    handler = _close_entry(_cb_update("opn:c:12"))
    check(
        getattr(handler, "callback", None) is journal.open_trades_close_entry,
        "🏁 detail button enters the close conversation (entry point)",
    )
    handler = _close_entry(_text_update("/close 12"))
    check(
        getattr(handler, "callback", None) is journal.close_start_text,
        "/close <id> enters the close conversation",
    )

    ckey = close_conv._get_key(_text_update("long"))

    def _routed_close(st, update):
        close_conv._conversations[ckey] = st
        result = close_conv.check_update(update)
        close_conv._conversations.clear()
        return result[2] if result else None

    close_cases = [
        (journal.CLOSE_STATUS, "✅ Win (TP)", journal.ask_close_status),
        (journal.CLOSE_DATE, "2026-03-02", journal.ask_close_date),
        (journal.CLOSE_HOUR, "15:45", journal.ask_close_hour),
        (journal.CLOSE_HOUR, "الان", journal.ask_close_hour),
        (journal.CLOSE_PRICE, "58500", journal.ask_close_price),
        (journal.CLOSE_PHOTOS, "-", journal.ask_close_photos_text),
        (journal.CLOSE_REASON, "target", journal.ask_close_reason),
        (journal.CLOSE_MOOD, "آرام", journal.ask_close_mood),
        (journal.CLOSE_CONFIRM, "✅ ثبت", journal.save_close_trade),
    ]
    for st, text, fn in close_cases:
        handler = _routed_close(st, _text_update(text))
        check(
            getattr(handler, "callback", None) is fn,
            f"close conv routes '{text}' to {fn.__name__}",
        )

    check(
        _routed_to(journal.ENTRY, _text_update("📊 آمار")) is None,
        "menu texts don't leak into conversation steps",
    )
    check(
        _routed_to(journal.NOTES, _text_update("🏠 منو")) is None,
        "menu taps don't leak into conversation answers",
    )
    check(
        _routed_to(journal.NOTES, _text_update("📥 اکسل")) is None,
        "export taps don't leak into conversation answers",
    )
    check(
        _menu_routes("✅ Win") == [] and _menu_routes("❌ Loss") == [],
        "result buttons are conversation answers, not menu labels",
    )
    handler = _routed_to(journal.RESULT, _text_update("➖ BE"))
    check(
        getattr(handler, "callback", None) is journal.ask_result,
        "BE button routes to ask_result",
    )

    print()
    if failures:
        print(f"{len(failures)} check(s) FAILED")
        return 1
    print("All checks passed.")
    return 0


sys.exit(asyncio.run(main()))
"""Spreadsheet export for the trading journal (/export command).

Builds a formatted .xlsx workbook (one row per trade) and returns a path
ready to be sent as a Telegram document. Export files land in exports/ next
to bot.py (override with the EXPORT_DIR environment variable) and are named
trades-YYYYMMDD-HHMMSS.xlsx.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

logger = logging.getLogger(__name__)

EXPORT_DIR = Path(
    os.getenv("EXPORT_DIR", str(Path(__file__).resolve().parent / "exports"))
)

_HEADERS = [
    "ID",
    "Symbol",
    "Market",
    "Direction",
    "Timeframe",
    "Entry",
    "Exit (hit)",
    "Take profit",
    "Stop loss",
    "Result",
    "Margin",
    "Risk %",
    "P&L",
    "ROI %",
    "Date",
    "Mood",
    "Notes",
    "Logged at",
]
_WIDTHS = [7, 12, 9, 10, 10, 10, 11, 11, 10, 6, 12, 8, 12, 8, 18, 12, 40, 19]


def build_workbook(path: Path, rows: list) -> None:
    """Write the trades workbook to *path* (xlsx via openpyxl)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col, (header, width) in enumerate(zip(_HEADERS, _WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    for r, row in enumerate(rows, start=2):
        values = [
            row["id"],
            row["symbol"],
            row["market"] or "",
            row["direction"].upper(),
            row["timeframe"] or "",
            row["entry_price"],
            row["exit_price"],
            row["take_profit"],
            row["stop_loss"],
            row["hit"] or "",
            row["size"],
            row["risk_percent"],
            row["pnl"],
            # Legacy rows have no stored ROI — compute it from P&L / margin.
            (
                row["roi"]
                if row["roi"] is not None
                else (row["pnl"] / row["size"] * 100 if row["size"] else None)
            ),
            row["trade_date"],
            row["mood"] or "",
            row["notes"],
            row["created_at"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (
                6, 7, 8, 9, 11, 12, 13, 14
            ):  # prices, margin, risk, P&L, ROI
                cell.number_format = "0.####"
            if col == 15:
                cell.number_format = "@"

    wb.save(path)
    logger.info("Exported %d trade(s) to %s", len(rows), path)


def build_export_file() -> Path:
    """Create a timestamped export workbook and return its path."""
    rows = db.get_all_trades()
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = EXPORT_DIR / f"trades-{stamp}.xlsx"
    build_workbook(path, rows)
    return path